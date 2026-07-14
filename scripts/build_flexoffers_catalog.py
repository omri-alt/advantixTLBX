#!/usr/bin/env python3
"""
Build ``data/flexoffers_advertisers.json`` from a FlexOffers AdvertiserData export.

  python scripts/build_flexoffers_catalog.py \"C:\\Users\\Acer\\Downloads\\AdvertiserData (1).xlsx\"
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

DEFAULT_OUT = ROOT / "data" / "flexoffers_advertisers.json"

SKIP_STATUS = {"declined", "deactivated", "null", ""}
STATUS_MAP = {
    "not applied for": "available",
    "approved": "approved",
    "pending": "pending",
}
GEO_ALIASES = {"gb": "uk"}


def _host_of(url: str) -> str:
    u = (url or "").strip()
    if not u:
        return ""
    if "://" not in u:
        u = "https://" + u
    try:
        h = (urlparse(u).hostname or "").lower().strip(".")
    except Exception:
        return ""
    if h.startswith("www."):
        h = h[4:]
    return h


def _norm_geo(country: object) -> str:
    s = str(country or "").strip().lower()
    if not s or s in ("null", "none", "n/a"):
        return ""
    s = GEO_ALIASES.get(s, s)
    if len(s) == 2 and s.isalpha():
        return s
    return s[:2] if len(s) >= 2 else ""


def build_catalog(xlsx_path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["AdvertiserDetails"] if "AdvertiserDetails" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}
    required = ("ID", "Name", "DomainURL", "Country", "Status", "AllowDeepLinking")
    missing = [c for c in required if c not in idx]
    if missing:
        raise SystemExit(f"Missing columns in {xlsx_path.name}: {missing}")

    advertisers: list[dict] = []
    skipped = 0
    for r in rows:
        status_raw = str(r[idx["Status"]] or "").strip()
        st_key = status_raw.lower()
        if st_key in SKIP_STATUS or st_key not in STATUS_MAP:
            skipped += 1
            continue
        url = str(r[idx["DomainURL"]] or "").strip()
        host = _host_of(url)
        if not host or "." not in host:
            skipped += 1
            continue
        try:
            aid = int(r[idx["ID"]])
        except (TypeError, ValueError):
            skipped += 1
            continue
        advertisers.append(
            {
                "id": aid,
                "name": str(r[idx["Name"]] or "").strip(),
                "host": host,
                "url": url if url.startswith("http") else f"https://{url}",
                "geo": _norm_geo(r[idx["Country"]]),
                "status": STATUS_MAP[st_key],
                "deeplink": bool(r[idx["AllowDeepLinking"]]),
            }
        )
    wb.close()

    by_id: dict[int, dict] = {}
    rank = {"approved": 3, "pending": 2, "available": 1}
    for a in advertisers:
        prev = by_id.get(a["id"])
        if not prev or rank.get(a["status"], 0) > rank.get(prev["status"], 0):
            by_id[a["id"]] = a

    clean = sorted(by_id.values(), key=lambda x: (x["host"], x["geo"], x["id"]))
    return {
        "updated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": xlsx_path.name,
        "count": len(clean),
        "skipped": skipped,
        "status_counts": dict(Counter(a["status"] for a in clean)),
        "advertisers": clean,
    }


def main() -> int:
    p = argparse.ArgumentParser(description="Build FlexOffers advertiser catalog JSON.")
    p.add_argument("xlsx", type=Path, help="Path to AdvertiserData.xlsx")
    p.add_argument(
        "--out",
        type=Path,
        default=DEFAULT_OUT,
        help=f"Output JSON (default: {DEFAULT_OUT})",
    )
    args = p.parse_args()
    if not args.xlsx.exists():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        return 1
    payload = build_catalog(args.xlsx)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    # Keep regenerator metadata out of the committed lookup file shape expected by flexoffers.py
    out_payload = {
        "updated_at": payload["updated_at"],
        "source": payload["source"],
        "count": payload["count"],
        "advertisers": payload["advertisers"],
    }
    args.out.write_text(
        json.dumps(out_payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(
        f"Wrote {args.out} advertisers={payload['count']} "
        f"skipped={payload['skipped']} statuses={payload['status_counts']} "
        f"bytes={args.out.stat().st_size}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
